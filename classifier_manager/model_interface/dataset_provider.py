import openpyxl

def read_xlsx1 (path,sheet_name, level = 1):
    column = 2
    if(level == 3):
        column = 4
    if(level == 2):
        column = 3
    le = openpyxl.load_workbook(path)
    sheet = le.get_sheet_by_name(sheet_name)
    descriptions = [sheet.cell(row = i, column = 1).value for i in range(2,sheet.max_row+1)]
    items = [sheet.cell(row = i, column = column).value for i in range(2,sheet.max_row+1)]
    return descriptions,items

def read_xlsx (path,sheet_name, level = 1):
    column = 6
    if(level == 3):
        column = 8 
    if(level == 2):
        column = 7
    le = openpyxl.load_workbook(path, data_only=True)
    sheet = le.get_sheet_by_name(sheet_name)
    descriptions = [sheet.cell(row = i, column = 2).value for i in range(2,sheet.max_row+1)]
    items = [sheet.cell(row = i, column = column).value for i in range(2,sheet.max_row+1)]
    return descriptions,items

def read_taxonomies(path, sheet_name, gold_column):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.get_sheet_by_name(sheet_name)
    banks = [ (i,sheet.cell(row = 1, column = i).value) for i in range(1, gold_column+1)]
    dictionary = {}
    for i in range (2, sheet.max_row+1):
        gold_el = sheet.cell(row = i, column = gold_column).value
        dictionary[gold_el] = {}
        for col,name in banks[:gold_column-1]:
            dictionary[gold_el][name] = sheet.cell(row = i, column=col).value
    return dictionary


