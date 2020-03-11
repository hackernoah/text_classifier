from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.feature_selection import chi2
from sklearn.feature_selection import SelectKBest
import numpy as np
import openpyxl
from sklearn.svm import LinearSVC
from sklearn import metrics
from sklearn.utils.multiclass import unique_labels
import random
from pprint import pprint

class RandomClassifier:

    def __init__(self, descriptions, items, kfolds, dist = False):
        self.kfolds = kfolds
        self.use_distribution = dist
        self.__populate_records(descriptions, items)
        self.__create_target()

    def __populate_records(self, descriptions, items):
        self.records = []
        self.original_records = []
        for i in range(len(items)):
            item = items[i]
            desc = descriptions[i]
            if( isinstance(item, str) and isinstance(desc, str) and (desc != None) and (item != None) and (item != '#N/A') ):
                self.records.append((desc,item))
        random.shuffle(self.records)
        self.original_records = self.records.copy()

    def __create_target(self):
        self.target_values = {}
        self.target_names = list(set([record[1] for record in self.records]))
        for i in range(len(self.target_names)):
            self.target_values[self.target_names[i]] = i
            self.target_values[str(i)] = self.target_names[i]
    
    def kfold_cross_validation(self, filepath = None):
        export_excel = openpyxl.Workbook()
        indexes = [ fold*(len(self.records) // self.kfolds) for fold in list(range(self.kfolds)) ]
        indexes.append(len(self.records))
        results = []
        reports = []
        for i in range(self.kfolds):
            test_records = self.records[indexes[i]:indexes[i+1]]
            test_target = [self.target_values[record[1]] for record in test_records]
            predictions = self.predict(test_records)
            test_target_names = list(unique_labels([self.target_names[target] for target in test_target],[self.target_names[target] for target in predictions]))
            report = metrics.classification_report(test_target, predictions,target_names=test_target_names,output_dict=True)
            if(filepath):
                export_excel = self.create_export_sheet(export_excel,self.original_records[indexes[i]:indexes[i+1]],test_records, predictions,report, "iteration {}".format(i+1))
            results.append((report['accuracy'], report['macro avg']))
            reports.append(report)
        if(filepath):
            sheet = export_excel.get_sheet_by_name("Sheet")
            export_excel.remove_sheet(sheet)
            export_excel = self.create_export_sheet(export_excel, None,None,None,self.get_reports_mean(reports), 'Media risultati')
            export_excel.save(filepath)
        return results

    def predict(self, records):
        predictions = []
        if(self.use_distribution):
            dist = self.get_distribution()
            prob = []
            for key in dist:
                p = round(dist[key]/len(self.records),4)
                prob.append(p)
            print(prob)
            for record in records:
                predictions.append(np.random.choice(np.arange(0, len(self.target_names)), p=prob))
        else:
            for record in records:
                predictions.append(random.randint(0,len(self.target_names) - 1 ))
        return predictions

    def get_distribution(self):
        labels = set([l for (d,l) in self.records])
        dist = {}
        for label in labels:
            dist[label] = 0
        for d,l in self.records:
            dist[l] += 1
        return dist

    def create_export_sheet(self,workbook, records, features, predictions, results, sheet_name):
        ws = workbook.create_sheet(title=sheet_name)
        if(records):
            ws.cell(column=1, row = 1, value="Descrizione originale")
            ws.cell(column=2, row = 1, value="Risultato preprocessamento")
            ws.cell(column=3, row = 1, value="Label originale")
            ws.cell(column=4, row = 1, value="Classificazione")
            for row in range(len(records)):
                ws.cell(column=1, row = row+2, value= records[row][0])
                ws.cell(column=2, row = row+2, value= features[row][0])
                ws.cell(column=3, row = row+2, value= records[row][1])
                ws.cell(column=4, row = row+2, value= self.target_names[predictions[row]])
        if(results):
            ws.cell(column=7, row = 1, value="Precision")
            ws.cell(column=8, row = 1, value="Recall")
            ws.cell(column=9, row = 1, value="F1-score")
            ws.cell(column=10, row = 1, value="Numero istanze")
            row = 2
            for key in results:
                ws.cell(column=6, row = row, value=str(key))
                if(key == 'accuracy'):
                    ws.cell(column=9, row = row, value=results[key])
                else:
                    ws.cell(column=7, row = row, value=results[key]['precision'])
                    ws.cell(column=8, row = row, value=results[key]['recall'])
                    ws.cell(column=9, row = row, value=results[key]['f1-score'])
                    ws.cell(column=10, row = row, value=results[key]['support'])
                    row += 1
        return workbook

    def get_reports_mean(self,results):
        new_results = results[0].copy()
        for key in new_results:
            prec, rec, fscore = [0.0 for _ in range(3)]
            count = 0
            support = 0
            for report in results:
                if(key == 'accuracy'):
                    prec += report[key]
                    count += 1
                else:
                    if(key in report.keys()):
                        prec += report[key]['precision']
                        rec += report[key]['recall']
                        fscore += report[key]['f1-score']
                        support += report[key]['support']
                        count += 1
            if(key == 'accuracy'):
                new_results[key] = prec / count
            elif(count > 0):
                new_results[key]['precision'] = prec / count
                new_results[key]['recall'] = rec / count
                new_results[key]['f1-score'] = fscore / count
                new_results[key]['support'] = support / count
            prec, rec, fscore = [0.0 for _ in range(3)]
            count = 0
            support = 0
        return new_results

    
    
